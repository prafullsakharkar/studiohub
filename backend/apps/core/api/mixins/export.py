"""
Export Operations Mixin.

This mixin provides export functionality for ViewSets.
"""
from __future__ import annotations

from typing import TYPE_CHECKING

from rest_framework import serializers
from rest_framework.response import Response

from apps.core.api.builders.export import ExportBuilder
from apps.core.api.utils.response import ResponseUtils

if TYPE_CHECKING:
    from django.db.models import QuerySet


class ExportMixin:
    """
    Mixin for export operations.
    """

    export_serializer_class: type[serializers.Serializer] | None = None
    export_format: str = "csv"  # Default export format

    def get_export_serializer(self) -> type[serializers.Serializer]:
        """Get the serializer class for export."""
        if self.export_serializer_class is None:
            raise NotImplementedError("export_serializer_class must be defined.")
        return self.export_serializer_class

    def export_data(self, request, *args, **kwargs) -> Response:
        """Handle export operations."""
        # Get filtered queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Serialize data
        serializer_class = self.get_export_serializer()
        serializer = serializer_class(queryset, many=True)

        # Build export metadata
        export_info = ExportBuilder.build(
            filename=self._get_export_filename(),
            content_type=self._get_export_content_type(),
        )

        # Return response based on format
        format_type = request.query_params.get("format", self.export_format)

        if format_type == "csv":
            return self._export_csv(serializer.data, export_info)
        elif format_type == "json":
            return self._export_json(serializer.data, export_info)
        elif format_type == "xlsx":
            return self._export_xlsx(serializer.data, export_info)
        else:
            return ResponseUtils.error(
                message="Unsupported export format.",
                code="unsupported_format",
                status_code=400,
            )

    def _get_export_filename(self) -> str:
        """Generate export filename."""
        model_name = self.serializer_class().Meta.model.__name__.lower()
        return f"{model_name}_export_{self._get_timestamp()}.csv"

    def _get_export_content_type(self) -> str:
        """Get export content type."""
        format_type = self.export_format
        if format_type == "csv":
            return "text/csv"
        elif format_type == "json":
            return "application/json"
        elif format_type == "xlsx":
            return (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        return "application/octet-stream"

    def _get_timestamp(self) -> str:
        """Get current timestamp for filename."""
        from datetime import datetime

        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def _export_csv(self, data: list[dict], export_info: dict) -> Response:
        """Export data as CSV."""
        import csv
        import io

        if not data:
            return ResponseUtils.error(
                message="No data to export.",
                code="no_data",
                status_code=404,
            )

        # Get headers from first record
        headers = list(data[0].keys())

        # Create CSV in memory
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)

        # Build response
        response = Response(output.getvalue())
        response["Content-Disposition"] = f'attachment; filename="{export_info["filename"]}"'
        response["Content-Type"] = export_info["content_type"]

        return response

    def _export_json(self, data: list[dict], export_info: dict) -> Response:
        """Export data as JSON."""
        import json

        if not data:
            return ResponseUtils.error(
                message="No data to export.",
                code="no_data",
                status_code=404,
            )

        response = Response(json.dumps(data, indent=2))
        response["Content-Disposition"] = f'attachment; filename="{export_info["filename"]}"'
        response["Content-Type"] = export_info["content_type"]

        return response

    def _export_xlsx(self, data: list[dict], export_info: dict) -> Response:
        """Export data as XLSX (Excel)."""
        try:
            import openpyxl
        except ImportError:
            return ResponseUtils.error(
                message="openpyxl is not installed. Install with: pip install openpyxl",
                code="missing_dependency",
                status_code=500,
            )

        if not data:
            return ResponseUtils.error(
                message="No data to export.",
                code="no_data",
                status_code=404,
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Save to bytes
        from io import BytesIO

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Build response
        response = Response(output.read())
        response["Content-Disposition"] = f'attachment; filename="{export_info["filename"]}"'
        response["Content-Type"] = export_info["content_type"]

        return response
